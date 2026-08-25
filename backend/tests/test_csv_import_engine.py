import io

import openpyxl
import pytest

from app.engines.csv_import import parse_holdings_csv, parse_holdings_file, parse_holdings_rows

_STANDARD_HEADER = "Instrument,Qty.,Avg. cost,LTP,Cur. val,P&L,Net chg.,Day chg."


def _xlsx_bytes(rows: list[list[object]], *, sheet_names: list[str] | None = None) -> bytes:
    """Builds a minimal in-memory .xlsx for tests — no real Upstox export
    available, so this stands in for one. `sheet_names` lets a test put
    extra (empty) sheets before/after the real data to exercise the
    active-sheet-isn't-the-holdings-sheet fallback."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (sheet_names or ["Sheet1"])[0]
    for row in rows:
        ws.append(row)
    for name in (sheet_names or [])[1:]:
        wb.create_sheet(name)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_standard_zerodha_headers_parse() -> None:
    csv_text = _STANDARD_HEADER + "\nTCS,10,3500.50,3600,36000,995,2.8,0.5\n"
    result = parse_holdings_csv(csv_text)
    assert result.errors == []
    assert len(result.rows) == 1
    row = result.rows[0]
    assert row.symbol == "TCS"
    assert row.quantity == 10
    assert row.avg_cost == 3500.50


def test_upstox_api_shaped_headers_resolve() -> None:
    # Upstox's own API field names (tradingsymbol/quantity/average_price,
    # per upstox/upstox-python's docs) — no confirmed literal export
    # header, but these are the best signal available and the tolerant
    # matching should resolve them without any Upstox-specific code.
    csv_text = "tradingsymbol,quantity,average_price,isin\nINFY,5,1500,INE009A01021\n"
    result = parse_holdings_csv(csv_text)
    assert result.errors == []
    assert result.rows[0].symbol == "INFY"
    assert result.rows[0].quantity == 5
    assert result.rows[0].avg_cost == 1500


def test_alternate_header_names_resolve() -> None:
    csv_text = "Tradingsymbol,Quantity,Average Price\nINFY,5,1500\n"
    result = parse_holdings_csv(csv_text)
    assert result.errors == []
    assert result.rows[0].symbol == "INFY"
    assert result.rows[0].quantity == 5
    assert result.rows[0].avg_cost == 1500


def test_header_matching_is_case_and_whitespace_tolerant() -> None:
    csv_text = "  INSTRUMENT  , QTY , AVG COST \nHDFC,1,100\n"
    result = parse_holdings_csv(csv_text)
    assert result.errors == []
    assert result.rows[0].symbol == "HDFC"


def test_empty_file_raises_value_error() -> None:
    with pytest.raises(ValueError, match="empty"):
        parse_holdings_csv("")


def test_missing_quantity_column_raises() -> None:
    with pytest.raises(ValueError, match="symbol, a quantity, an average cost"):
        parse_holdings_csv("Instrument,Avg. cost\nTCS,100\n")


def test_leading_title_rows_are_skipped_to_find_the_real_header() -> None:
    # A common real-world report shape: a title/"generated on"/account-id
    # row (or two) before the actual column header row.
    csv_text = (
        "Holdings Report\n"
        "Generated on 25-Aug-2026 for account XX1234\n"
        + _STANDARD_HEADER
        + "\nTCS,10,3500,3600,36000,995,0,0\n"
    )
    result = parse_holdings_csv(csv_text)
    assert result.errors == []
    assert result.rows[0].symbol == "TCS"


def test_header_never_found_within_scan_window_raises() -> None:
    junk_rows = "\n".join(f"junk row {i}" for i in range(20))
    with pytest.raises(ValueError, match="couldn't find"):
        parse_holdings_csv(junk_rows)


def test_unparseable_number_becomes_row_error_not_a_crash() -> None:
    csv_text = (
        _STANDARD_HEADER
        + "\nTCS,not-a-number,3500,3600,36000,0,0,0\nINFY,5,1500,1600,8000,500,0,0\n"
    )
    result = parse_holdings_csv(csv_text)
    assert len(result.rows) == 1
    assert result.rows[0].symbol == "INFY"
    assert len(result.errors) == 1
    assert result.errors[0].row_number == 1
    assert "unparseable" in result.errors[0].reason


def test_zero_or_negative_quantity_becomes_row_error() -> None:
    csv_text = (
        _STANDARD_HEADER + "\nTCS,0,3500,0,0,0,0,0\nINFY,-5,1500,0,0,0,0,0\nHDFC,5,1500,0,0,0,0,0\n"
    )
    result = parse_holdings_csv(csv_text)
    assert [r.symbol for r in result.rows] == ["HDFC"]
    assert len(result.errors) == 2
    assert all("positive" in e.reason for e in result.errors)


def test_zero_or_negative_avg_cost_becomes_row_error() -> None:
    csv_text = _STANDARD_HEADER + "\nTCS,10,0,0,0,0,0,0\n"
    result = parse_holdings_csv(csv_text)
    assert result.rows == []
    assert "positive" in result.errors[0].reason


def test_blank_trailing_row_is_silently_skipped() -> None:
    csv_text = _STANDARD_HEADER + "\nTCS,10,3500,3600,36000,995,0,0\n,,,,,,,\n"
    result = parse_holdings_csv(csv_text)
    assert len(result.rows) == 1
    assert result.errors == []


def test_empty_symbol_cell_becomes_row_error() -> None:
    csv_text = _STANDARD_HEADER + "\n,10,3500,3600,36000,995,0,0\n"
    result = parse_holdings_csv(csv_text)
    assert result.rows == []
    assert len(result.errors) == 1
    assert "empty symbol" in result.errors[0].reason


def test_duplicate_symbol_keeps_first_errors_second() -> None:
    csv_text = (
        _STANDARD_HEADER
        + "\nTCS,10,3500,3600,36000,995,0,0\nTCS,20,3600,3600,72000,0,0,0\n"
    )
    result = parse_holdings_csv(csv_text)
    assert len(result.rows) == 1
    assert result.rows[0].quantity == 10
    assert len(result.errors) == 1
    assert "duplicate symbol TCS" in result.errors[0].reason
    assert "row 1" in result.errors[0].reason


def test_comma_thousands_separators_parse() -> None:
    csv_text = _STANDARD_HEADER + '\nTCS,"1,234.5","3,500.75",0,0,0,0,0\n'
    result = parse_holdings_csv(csv_text)
    assert result.errors == []
    assert result.rows[0].quantity == 1234.5
    assert result.rows[0].avg_cost == 3500.75


def test_pnl_and_current_value_columns_are_never_consulted() -> None:
    csv_text = "Instrument,Qty.,Avg. cost,P&L,Cur. val\nTCS,10,3500,garbage,also garbage\n"
    result = parse_holdings_csv(csv_text)
    assert result.errors == []
    assert result.rows[0].symbol == "TCS"


def test_parse_holdings_rows_used_directly() -> None:
    rows = [["Instrument", "Qty.", "Avg. cost"], ["TCS", "10", "3500"]]
    result = parse_holdings_rows(rows)
    assert result.rows[0].symbol == "TCS"


def test_xlsx_parses_via_dispatch() -> None:
    raw = _xlsx_bytes([["Instrument", "Qty.", "Avg. cost"], ["TCS", 10, 3500]])
    result = parse_holdings_file(raw, "holdings.xlsx")
    assert result.errors == []
    assert result.rows[0].symbol == "TCS"
    assert result.rows[0].quantity == 10
    assert result.rows[0].avg_cost == 3500


def test_xlsx_with_leading_title_row_finds_the_real_header() -> None:
    raw = _xlsx_bytes(
        [
            ["Holdings Report — generated 25-Aug-2026"],
            ["Instrument", "Qty.", "Avg. cost"],
            ["INFY", 5, 1500],
        ]
    )
    result = parse_holdings_file(raw, "holdings.xlsx")
    assert result.errors == []
    assert result.rows[0].symbol == "INFY"


def test_xlsx_falls_back_to_a_non_active_sheet() -> None:
    wb = openpyxl.Workbook()
    summary_sheet = wb.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Not the holdings table"])
    holdings_sheet = wb.create_sheet("Holdings")
    holdings_sheet.append(["Instrument", "Qty.", "Avg. cost"])
    holdings_sheet.append(["HDFC", 3, 1600])
    buf = io.BytesIO()
    wb.save(buf)

    result = parse_holdings_file(buf.getvalue(), "holdings.xlsx")
    assert result.errors == []
    assert result.rows[0].symbol == "HDFC"


def test_xlsx_dispatch_is_case_insensitive_on_extension() -> None:
    raw = _xlsx_bytes([["Instrument", "Qty.", "Avg. cost"], ["TCS", 10, 3500]])
    result = parse_holdings_file(raw, "Holdings.XLSX")
    assert result.rows[0].symbol == "TCS"


def test_csv_dispatch_for_non_xlsx_filename() -> None:
    raw = (_STANDARD_HEADER + "\nTCS,10,3500,3600,36000,995,0,0\n").encode("utf-8")
    result = parse_holdings_file(raw, "holdings.csv")
    assert result.rows[0].symbol == "TCS"


def test_corrupt_xlsx_raises_value_error_not_a_crash() -> None:
    with pytest.raises(ValueError, match="Excel"):
        parse_holdings_file(b"not a real xlsx file", "holdings.xlsx")
